import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import os
import carregarDados as CDD
import salvarDados as SDD
import calendar
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
                 
data = datetime.now()
DATA = data.strftime("%d/%m/%Y") #data atual DD/MM/AAAA
BORDAFINA = Border(left=Side(style="thin"),
                   right=Side(style="thin"),
                   top=Side(style="thin"),
                   bottom=Side(style="thin"))

#função para fechar todas as janelas filhas da var-janela
def janelasKill(janela):
    for widget in janela.winfo_children():
        if isinstance(widget, tk.Toplevel):
            widget.destroy()

#função para fechar janelas
def cancelar(janela):
    janela.destroy()

#função para dar focus ao uma janela
def focus(janela):
    janela.focus_force()

#função para entrada de mercadorias de insumo, NFs ou cupons fiscais e suas informações
def NFInsumo(janela):
    janelasKill(janela=janela)
    try:
        DB = CDD.carregarArquivo("recebimentos.txt")
    except:
        DB = {DATA: {"notas": [], "cupons": []}}
        SDD.salvarArquivo("recebimentos.txt", DB)
        messagebox.showinfo('Arquivo DB','Arquivo de dados inexistente! Novo arquivo de dados de recebimentos criado com sucesso!')
        DB = CDD.carregarArquivo("recebimentos.txt")

    #função que salva as informações do recebimento dos insumos
    def salvar():
        if dataRecebimento.get().strip()=="" and recebedor.get().strip() == "" and nota.get().strip() =="" and cupom.get() == False and produtos.get("1.0", "end").strip() == "" and valor.get().strip() =="":
            return
        DATARECEBIMENTO = dataRecebimento.get().strip()
        RECEBEDOR = recebedor.get().strip()
        NUMDOC = nota.get().strip()
        if DATARECEBIMENTO == "":
            DATARECEBIMENTO = DATA
        else:
            try:
                datetime.strptime(DATARECEBIMENTO, "%d/%m/%Y")
            except:
                messagebox.showerror("Data", "Data digitada inválida ou fora do padrão DD/MM/AAAA")
                dataRecebimento.delete(0, tk.END)
                dataRecebimento.focus_set()
                return

        if RECEBEDOR == "":
            messagebox.showerror("ERRO", "Informe o nome do recebedor da mercadoria!")
            recebedor.focus_set()
            return
        else:
            RECEBEDOR = RECEBEDOR.lower().title()

        if NUMDOC == "":
            messagebox.showerror("ERRO", "Dígite o Numero do Documento!")
            nota.focus_set()
            return
        elif not cupom.get():
            try:
                NUMDOC=int(NUMDOC)
                NUMDOC=str(NUMDOC)
            except:
                messagebox.showerror("ERRO", "Digite apenas numeros inteiros para Notas Fiscais!")
                nota.delete(0, tk.END)
                nota.focus_set()
                return

        if not cupom.get():
            if DATARECEBIMENTO not in DB:
                DB[DATARECEBIMENTO] = {"notas": [], "cupons": []}

            DB[DATARECEBIMENTO]["notas"].append({"nota": NUMDOC, "recebedor": RECEBEDOR})

        else:
            PRODUTOS = produtos.get("1.0", "end").strip()
            VALORTOTAL = valor.get().strip()
            if PRODUTOS == "":
                messagebox.showerror("ERRO", "Descrição dos produtos é obrigatorio para 'Cupom Fiscal'!")
                produtos.focus_set()
                return
            
            if VALORTOTAL == "":
                messagebox.showerror("ERRO", "Valor Total é obrigatorio para 'Cupom Fiscal'!")
                valor.focus_set()
                return
            
            try:
                VALORTOTAL = float(VALORTOTAL.replace(",", "."))
                VALORTOTAL = str(VALORTOTAL)
            except:
                messagebox.showerror("ERRO", "Dígite Apenas Numeros no Campo 'Valor Total'!")
                valor.delete(0, tk.END)
                valor.focus_set()
                return
                
            if DATARECEBIMENTO not in DB:
                DB[dataRecebimento] = {"notas": [], "cupons": []}

            DB[DATARECEBIMENTO]["cupons"].append({"cupom": NUMDOC, "descricao": PRODUTOS, "valor": VALORTOTAL, "recebedor": RECEBEDOR})

        SDD.salvarArquivo("recebimentos.txt", DB)
        messagebox.showinfo("Sucesso", "Recebimento incluido com sucesso!")
        NFInsumo(janela=janela)

    #função para quando apertar enter na digitação da descrição dos produtos nao chamar o botaoConfirmar
    def enterHandler(event=None):
        widget_focado = janela.focus_get()
        if widget_focado == produtos:
            return  # deixa o Text agir normalmente (adiciona nova linha)
        botaoConfirmar.invoke()

    #função para quando apertar TAB na digitação da descrição dos produtos o cod pular pro proximo campo de preenchimento
    def tabHandler(event):
        event.widget.tk_focusNext().focus()
        return "break"  # impede o tab de ser inserido no Text

    #janela de recebimento de insumos
    janelaRecebimento= tk.Toplevel(janela)
    janelaRecebimento.title("Recebimento de Mercadoria")

    tk.Label(janelaRecebimento, text="Data").grid(column=0, row=0)
    dataRecebimento = tk.Entry(janelaRecebimento, width=15)
    dataRecebimento.grid(column=0, row=1, pady=(0,10))

    tk.Label(janelaRecebimento, text="Recebedor").grid(column=1, row=0)
    recebedor = tk.Entry(janelaRecebimento, width=15)
    recebedor.grid(column=1, row=1, pady=(0,10))

    tk.Label(janelaRecebimento, text="Numero do Documento\n(Nota Fiscal ou Cupom Fiscal)").grid(column=0, row=2, columnspan=2)
    nota = tk.Entry(janelaRecebimento, width=20)
    nota.grid(column=0, row=3, columnspan=2)

    cupom = tk.BooleanVar()
    checkCupom = tk.Checkbutton(janelaRecebimento, text="Cupom Fiscal", variable= cupom)
    checkCupom.grid(row=4, column=0, columnspan=2, pady=(0,10))
    
    tk.Label(janelaRecebimento, text="Descrição dos Produtos e Quantidades").grid(column=0, row=5, columnspan=2)
    produtos = tk.Text(janelaRecebimento, width=37, height=7)
    produtos.grid(column=0, row=6, columnspan=2, pady=(0,10), padx=20)

    produtos.bind("<Tab>", tabHandler)

    tk.Label(janelaRecebimento, text="Valor Total: ").grid(column=0, row=7, columnspan=2, pady=(0,10), padx=(42,0), sticky="w")
    valor = tk.Entry(janelaRecebimento, width=15)
    valor.grid(column=0, row=7, columnspan=2, pady=(0,10), padx=(0,42), sticky="e")

    botaoConfirmar = tk.Button(janelaRecebimento, text="Confirmar", padx=15, command= salvar)
    botaoConfirmar.grid(column=0, row=8, pady=10)

    botaoCancelar = tk.Button(janelaRecebimento, text="Cancelar", padx=15, command= lambda: cancelar(janela= janelaRecebimento))
    botaoCancelar.grid(column=1, row=8, pady=10)

    janelaRecebimento.bind('<Return>', enterHandler)
    recebedor.focus()

#função para criar um excel com o relatorio de gastos mensais da empresa baseado em cupons fiscais, e um relatorio de NFs recebidas.
def relatorioGastosMensais(janela):
    janelasKill(janela=janela)
    try:
        DB = CDD.carregarArquivo("recebimentos.txt")
    except:
        messagebox.showerror('ERRO','Arquivo de dados inexistente!')
        return

    #função que gera todas as datas possiveis do mes baseado no dia de entrada do usuario
    def gerarTodasDatasMes(mes, ano):
        totalDias = calendar.monthrange(ano, mes)[1]  # Retorna (dia_da_semana_inicio, total_de_dias)
        return [f"{dia:02d}/{mes:02d}/{ano}" for dia in range(1, totalDias + 1)]

    #função que cria e abre o excel
    def gerarExcel(DB, DATAS):
        VALORTOTAL = 0.0
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos Mensais"

        ws.append(["Data", "Código do Cupom Fiscal", "Descrição", "Valor", "Recebedor/ Comprador"])

        for data in DATAS:
            registros = DB.get(data, {})
            
            for cupom in registros.get("cupons", []):
                try:
                    VALORTOTAL += float(cupom.get("valor", "0").replace(",", "."))
                except:
                    pass
                ws.append([
                    data,
                    cupom.get("cupom", ""),
                    cupom.get("descricao", ""),
                    cupom.get("valor", ""),
                    cupom.get("recebedor", "")
                ])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 80))

        ultimaLinha = ws.max_row

        for row in ws.iter_rows(min_row=1, max_row=ultimaLinha, min_col=1, max_col=5):
            for cell in row:
                cell.border = BORDAFINA

        for row in range(1, ws.max_row + 1):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center")


        ws[f"D{ultimaLinha+1}"] = f"Valor Total: {VALORTOTAL:.2f}"
        ws.merge_cells(start_row=ultimaLinha+1, start_column=4, end_row=ultimaLinha+1, end_column=5)
        for row in ws.iter_rows(min_row=ultimaLinha+1, max_row=ultimaLinha + 1, min_col=4, max_col=5):
            for cell in row:
                cell.border = BORDAFINA

        linhaNotas = ultimaLinha + 3
        ws[f"A{linhaNotas}"] = "Notas Fiscais:"
        ws.merge_cells(start_row=linhaNotas, start_column=1, end_row=linhaNotas, end_column=5)
        cell = ws.cell(row=linhaNotas, column=1)
        cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=linhaNotas, max_row=linhaNotas, min_col=1, max_col=5):
            for cell in row:
                cell.border = BORDAFINA

        col = 1
        maxCols = 5
        row = linhaNotas + 1

        todasNotas = []
        for data in DATAS:
            for nota in DB.get(data, {}).get("notas", []):
                todasNotas.append(nota.get("nota", ""))

        for i, nota in enumerate(todasNotas):
            colAtual = (i % maxCols) + 1  # colunas de 1 a 5
            rowAtual = row + (i // maxCols)
            cell = ws.cell(row=rowAtual, column=colAtual, value=nota)
            cell.border = BORDAFINA
            cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save("relatorio de gastos mensais.xlsx")
        os.startfile("relatorio de gastos mensais.xlsx")

    #função para procurar as datas no DB
    def procurar(DATABUSCA):
        if not mesAtual.get():
            if DATABUSCA == "":
                return
            try:
                DATABUSCA=datetime.strptime(DATABUSCA, "%d/%m/%Y")
            except:
                messagebox.showerror("Data", "Data digitada inválida ou fora do padrão DD/MM/AAAA")
                inputData.delete(0, tk.END)
                inputData.focus_set()
                return

        else:
            DATABUSCA = datetime.strptime(DATA, "%d/%m/%Y")

        MES = DATABUSCA.month
        ANO = DATABUSCA.year
        datasMes = gerarTodasDatasMes(MES, ANO)
        DATASENCONTRADAS = [d for d in datasMes if d in DB]

        if not DATASENCONTRADAS:
            messagebox.showerror("Data", f"Nenhum dado encontrado para {MES:02d}/{ANO}")
            inputData.delete(0, tk.END)
            inputData.focus_set()
            return

        gerarExcel(DB=DB, DATAS=DATASENCONTRADAS)

    #janela para buscar a data do relatorio
    janelaBuscaData=tk.Toplevel(janela)
    janelaBuscaData.title('Relatorio de Gastos Mensais')

    tk.Label(janelaBuscaData,text='Data').grid(row=0,column=1,padx=5,pady=5)
    inputData=tk.Entry(janelaBuscaData, width=15)
    inputData.grid(row=1,column=1,padx=5,pady=5)
    inputData.focus_set()

    tk.Label(janelaBuscaData, text="DD/MM/AAAA").grid(row=2, column=1)

    mesAtual = tk.BooleanVar()
    checkMesAtual = tk.Checkbutton(janelaBuscaData, text="Mes Atual", variable=mesAtual)
    checkMesAtual.grid(row=3, column=1, padx=5, pady=5)

    botaoProcurar= tk.Button(janelaBuscaData, text='Procurar', command=lambda: procurar(inputData.get().strip()), default="active", padx=15)
    botaoProcurar.grid(row=5,column=0,pady=(5,10),padx=(10,20))

    botaoCancelar= tk.Button(janelaBuscaData, text='Cancelar', command=lambda:cancelar(janelaBuscaData), padx=15)
    botaoCancelar.grid(row=5, column=3, pady=(5,10),padx=(20,10)) 

    janelaBuscaData.bind('<Return>', lambda event=None: botaoProcurar.invoke())
