import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.cell.cell import MergedCell
import json
import os
from datetime import datetime, timedelta
import utils
import salvarDados as SDD
                 
data = datetime.now()
DATA = data.strftime("%d/%m/%Y")
DIASEMANA = data.weekday() #Dia atual DD/MM/AAAA
FONTENEGRITO = Font(bold=True)
BORDAFINA = Border(left=Side(style="thin"),
                   right=Side(style="thin"),
                   top=Side(style="thin"),
                   bottom=Side(style="thin"))
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
PASTAROMANEIOS = os.path.join(DESKTOP, "Romaneios de Expedição")
if not os.path.exists(PASTAROMANEIOS):
    os.makedirs(PASTAROMANEIOS)

#Função para carregar o arquivo(recebe como parametro o nome do arquivo a abrir). abre o arquivo e retorna ele.
def carregarArquivo(nomeArquivo=str):
    with open(nomeArquivo, 'r') as arquivo:
        conteudoArquivo=json.load(arquivo)
    return conteudoArquivo

#Função para abrir um excel com o romaneio de expedição da transportadora selecionada com as notas fiscais e informações referentes a elas
def romaneio(janela):
    utils.janelasKill(janela=janela)
    #cria o romaneio de expedição em um arquivo excel e o abre emseguida
    def criarRomaneio(pedidos):
        TRANSPORTADORA = pedidos[0]["transportadora"]
        TITULOSCOLUNAS = ["Nota", "Nome", "Cidade/ UF", "Volume", "Transportadora"]
        max_lengths = [len(title) for title in TITULOSCOLUNAS]  # lista com quantidade de letras de cada titulo em "TITULOCOLUNAS"
        VOLUMETOTAL = 0
        NOMEARQUIVO = f"romaneio de expedição {TRANSPORTADORA} {DATA.replace("/","-")}.xlsx"
        CAMINHOARQUIVO = os.path.join(PASTAROMANEIOS, NOMEARQUIVO)

        wb = Workbook()
        ws = wb.active
        ws.title = "Romaneio"

        ws["A1"] = f"Romaneio de Expedição"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        cell = ws.cell(row=1, column=1)
        cell.alignment = Alignment(horizontal="center")

        ws["A2"] = f"Transportadora {TRANSPORTADORA}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        cell = ws.cell(row=2, column=1)
        cell.alignment = Alignment(horizontal="center")

        ws["A3"] = f"Data: {DATA}"
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
        cell = ws.cell(row=3, column=1)
        cell.alignment = Alignment(horizontal="center")

        ws.append([])

        ws.append(TITULOSCOLUNAS)

        for pedido in pedidos:
            VOLUMETOTAL += int(pedido.get("vol.", "0"))

            linha = [
                pedido.get("nota"),
                ' '.join(pedido.get("nome").strip().lower().title().split()[:3]).lstrip('0123456789'),
                pedido.get("cidade/ uf"),
                pedido.get("vol."),
                pedido.get("transportadora")
            ]

            pedido["enviado"] = DATA

            ws.append(linha)

            for i, texto in enumerate(linha): # Atualiza os comprimentos máximos
                max_lengths[i] = max(max_lengths[i], len(str(texto)))

        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = BORDAFINA

        totalLinhas = ws.max_row + 1
        ws["D" + str(totalLinhas)] = f"Total de volumes: {VOLUMETOTAL}"
        ws.merge_cells(start_row=totalLinhas, start_column=4, end_row=totalLinhas, end_column=5)
        ws.cell(row=totalLinhas, column=4).alignment = Alignment(horizontal="right")

        ws.append([])
        ws.append(["Assinatura: "])
        ws[f"A{ws.max_row}"].font = FONTENEGRITO

        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)

        for i, length in enumerate(max_lengths):    # Ajustar largura das colunas
            col_letter = chr(65 + i) 
            ws.column_dimensions[col_letter].width = length + 2

        
        wb.save(CAMINHOARQUIVO)
        SDD.salvarArquivo("DB.txt", DB)
        os.startfile(CAMINHOARQUIVO)

    #cria o arquivo excel da relação de canhotos das nfs
    def relacaoCanhotosNfs(notas):
        TRANSPORTADORA = notas[0]["transportadora"]
        ARQUIVOORIGINAL = "relação de canhotos de notas fiscais.xlsx" #arquivo previamente criado que vai ser modificado
        CELULAS = ["A6", "A17", "A28", "A39"] #celulas que serao modificadas
        CELULASCHAVE = ["A8", "A19", "A30", "A41"] #celulas das chaves de nfs que serao modificadas
        MAXIMOPORFOLHA = len(CELULAS) # Quantidade de notas por aba
        NOVOTEXTOTRANS = f"Transportadora {TRANSPORTADORA}"
        NOVOTEXTODATA = f"Data {DATA}"
        NOMEARQUIVO = f"relação de canhotos de notas fiscais {TRANSPORTADORA} {DATA.replace("/","-")}.xlsx"
        CAMINHOARQUIVO = os.path.join(PASTAROMANEIOS, NOMEARQUIVO)

        wb = load_workbook(ARQUIVOORIGINAL)

        # Dividir as notas para diferentes abas
        for index in range(0, len(notas), MAXIMOPORFOLHA):
            parteListaNFs = notas[index: index + MAXIMOPORFOLHA]

            if index == 0:
                ws = wb.active  # Primeira folha
                ws.title = "relação de canhotos"
            else:
                ws = wb.copy_worksheet(wb.active)  # Cria uma cópia da aba original
                ws.title = f"relação de canhotos {index // MAXIMOPORFOLHA + 1}"  # Nomeia a aba sequencialmente

            ws["A2"] = NOVOTEXTOTRANS
            ws["A2"].font = FONTENEGRITO

            ws["D3"] = NOVOTEXTODATA
            ws["D3"].font = FONTENEGRITO

            for i in range(len(CELULASCHAVE)):
                if i < len(parteListaNFs):
                    celula = ws[CELULASCHAVE[i]]
                    text = ws[CELULASCHAVE[i]].value
                    if "chave NF" in parteListaNFs[i]:
                        text = text + " " + parteListaNFs[i]["chave NF"]
                        celula.value = text
                        celula.font = FONTENEGRITO
                    else:
                        ws[CELULASCHAVE[i]] = text
                        ws[CELULASCHAVE[i]].font = FONTENEGRITO
            
            # Preencher as células com os valores e limpar as restantes
            for i in range(len(CELULAS)):  # Itera sobre todas as células disponíveis na folha
                if i < len(parteListaNFs):  # Se ainda houver itens na lista, preenche normalmente
                    ws[CELULAS[i]] = f"Nota: {parteListaNFs[i]['nota']}, Nome: {' '.join(parteListaNFs[i]["nome"].strip().lower().title().split()[:5]).lstrip('0123456789')}, Cidade: {parteListaNFs[i]['cidade/ uf']}, {parteListaNFs[i]['vol.']} caixa{'s' if parteListaNFs[i]['vol.'] != '1' else ''}."
                    ws[CELULAS[i]].font = FONTENEGRITO
                    ws[CELULAS[i]].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                else:
                    # Limpa as células não utilizadas (caso a lista seja menor que o número de células disponíveis)
                    linha = int(coordinate_from_string(CELULAS[i])[1])
                    for row in ws.iter_rows(min_row=linha-1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            if not isinstance(cell, MergedCell):
                                cell.border = Border()
                                cell.value = None

        wb.save(CAMINHOARQUIVO)
        os.startfile(CAMINHOARQUIVO)

    def imprimirRomaneio(transportadora, janela):
        pedidosRomaneio=[]
        if not incluirRefrigeradas.get():
            for pedido in pedidosNaoEnviados:
                if pedido["transportadora"] == transportadora:
                    if pedido["refrigerado"] == "nao":
                        pedidosRomaneio.append(pedido)
        else:
            for pedido in pedidosNaoEnviados:
                if pedido["transportadora"] == transportadora:
                    pedidosRomaneio.append(pedido)

        if len(pedidosRomaneio) == 0:
            messagebox.showinfo("DB", "Sem pedidos não refigerados para a transportadora selecionada.")
            utils.focus(janelaSelecaoTrans)
            return

        criarRomaneio(pedidosRomaneio)
        if relacaoDeCanhotos.get():
            relacaoCanhotosNfs(pedidosRomaneio)

        utils.cancelar(janela=janela)
    
    try:
        DB=carregarArquivo('DB.txt')
    except:
        messagebox.showerror('ERRO','Arquivo de dados inexistente!')
        return
    
    pedidosNaoEnviados=[]
    for dia, listaPedidos in DB.items():
        for pedido in listaPedidos:
            if pedido["enviado"]=="nao":
                pedidosNaoEnviados.append(pedido)

    if pedidosNaoEnviados:
        transportadoras = []
        for pedido in pedidosNaoEnviados:
            if pedido["transportadora"] not in transportadoras:
                transportadoras.append(pedido["transportadora"])

        #janela de seleção da transportadora
        janelaSelecaoTrans=tk.Toplevel(janela)
        janelaSelecaoTrans.title('Romaneio')

        tk.Label(janelaSelecaoTrans,text='Selecione a Transportadora').grid(row=0,column=0,columnspan=2, padx=5,pady=(5,0))
        listbox = tk.Listbox(janelaSelecaoTrans, selectmode=tk.SINGLE, width=40, height=len(transportadoras))
        for transportadora in transportadoras:
            listbox.insert(tk.END, transportadora)
        listbox.grid(row=1,column=0,columnspan=2,padx=10,pady=(0,5))
        listbox.bind("<Double-1>", lambda event: imprimirRomaneio(listbox.get(listbox.curselection()), janelaSelecaoTrans))

        #relação de canhotos das notas fiscais em caso de entrega propria não ser preciso o destaque do canhoto da nota fiscal original
        relacaoDeCanhotos = tk.BooleanVar()
        check_relacao = tk.Checkbutton(janelaSelecaoTrans, text="Relação de Canhotos das Notas Fiscais", variable=relacaoDeCanhotos)
        check_relacao.grid(row=2, column=0, columnspan=2, padx=10, pady=(5,0), sticky="w")

        #incluir ou nao mercadoria refrigerada no romaneio de expedição 
        incluirRefrigeradas = tk.BooleanVar()
        incluirRefrigeradas.set(True)
        check_refrigeradas = tk.Checkbutton(janelaSelecaoTrans, text="Icluir Mercadorias Refrigeradas", variable=incluirRefrigeradas)
        check_refrigeradas.grid(row=3, column=0, columnspan=2, padx=10, pady=(0,5), sticky="w")

        botaoSelecionar= tk.Button(janelaSelecaoTrans, text='selecionar', command=lambda:imprimirRomaneio(listbox.get(listbox.curselection()), janelaSelecaoTrans), default="active", padx=11)
        botaoSelecionar.grid(row=4,column=0,pady=5,padx=5)

        botaoCancelar= tk.Button(janelaSelecaoTrans, text='Cancelar', command=lambda:utils.cancelar(janelaSelecaoTrans), padx=15)
        botaoCancelar.grid(row=4, column=1, pady=5,padx=5)
        
        janelaSelecaoTrans.bind('<Return>', lambda event=None: botaoSelecionar.invoke())

    else:
        messagebox.showerror('ERRO','Sem pedidos separados e não enviados!')
        return
    
#Faz buscas por pedidos ja separados inclusos no DB atravez do numero da nf ou do nome do cliente
def acharPedido(janela):
    utils.janelasKill(janela=janela)
    try:
        DB = carregarArquivo('DB.txt')
    except:
        messagebox.showerror('DB', 'Arquivo de dados inexistente!')
        return
    
    def busca():
        NOTAINPUT=inputNota.get().strip()
        NOMEINPUT=inputNome.get().strip().title()
        if NOTAINPUT == "" and NOMEINPUT == "":
            return
        
        #busca pelo nome
        elif NOTAINPUT == "" and NOMEINPUT != "":
            NOTASENCONTRADAS = []
            for dia, listaPedidos in DB.items():
                for pedido in listaPedidos:
                    if len(NOMEINPUT.split()) == 1:
                        if pedido["nome"].split()[0].title() == NOMEINPUT.split()[0].title():
                            pedido["separado"] = dia
                            NOTASENCONTRADAS.append(pedido)

                    if len(NOMEINPUT.split()) == 2 and len(pedido["nome"].split()) > 1:
                        if pedido["nome"].split()[0].title() == NOMEINPUT.split()[0].title() and pedido["nome"].split()[1].title() == NOMEINPUT.split()[1].title():
                            pedido["separado"] = dia
                            NOTASENCONTRADAS.append(pedido)

                    if len(NOMEINPUT.split()) >= 3 and len(pedido["nome"].split()) > 2:
                        if pedido["nome"].split()[0].title() == NOMEINPUT.split()[0].title() and pedido["nome"].split()[1].title() == NOMEINPUT.split()[1].title() and pedido["nome"].split()[2].title() == NOMEINPUT.split()[2].title():
                            pedido["separado"] = dia
                            NOTASENCONTRADAS.append(pedido)

            if len(NOTASENCONTRADAS) > 0:
                LINHA = 0
                def on_mouse_wheel(event):
                    canvas.yview_scroll(-1 * (event.delta // 120), "units")

                #janela que mostra as notas encontradas
                janelaResultado = tk.Toplevel(janela)
                janelaResultado.title('Notas Encontradas')

                frame = tk.Frame(janelaResultado)
                frame.grid(row=0, column=0, padx=5, pady=5, columnspan=2)

                tk.Label(frame, text="Notas Encontradas").grid(row=0, column=0, padx=5, pady=5, columnspan=2)

                canvas = tk.Canvas(frame, height=400)
                scrollbar = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
                scrollable_frame = tk.Frame(canvas)

                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(
                        scrollregion=canvas.bbox("all")
                    )
                )

                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)

                canvas.grid(row=1, column=0, sticky="nsew")
                scrollbar.grid(row=1, column=1, sticky="ns")

                janelaResultado.bind_all("<MouseWheel>", on_mouse_wheel)

                for nota in NOTASENCONTRADAS:
                    LINHA += 1
                    pedidoNota = nota["nota"]
                    pedidoNome = nota["nome"]
                    pedidoCidade = nota.get("cidade/ uf", " ")
                    pedidoVol = nota["vol."]
                    pedidoTrans = nota["transportadora"]
                    pedidoSep = nota.get("separador", " ")
                    pedidoRefri = nota.get("refrigerado", " ")
                    pedidoEnviado = nota["enviado"] 
                    dataSeparada = nota["separado"]

                    tk.Label(scrollable_frame, text=f"Nota: {pedidoNota}").grid(row=LINHA, column=0, padx=5, pady=5)
                    
                    listbox = tk.Listbox(scrollable_frame, height=7, width=40)
                    listbox.grid(row=LINHA, column=1, padx=5, pady=5)
                    listbox.insert(tk.END, f"Nome: {pedidoNome}")
                    listbox.insert(tk.END, f"Cidade/ UF: {pedidoCidade}")
                    listbox.insert(tk.END, f"Volumes: {pedidoVol}")
                    listbox.insert(tk.END, f"Refrigerado: {pedidoRefri}")
                    listbox.insert(tk.END, f"Transportadora: {pedidoTrans}")
                    listbox.insert(tk.END, f"Separado: {dataSeparada} por {pedidoSep}")
                    listbox.insert(tk.END, f"Enviado: {pedidoEnviado}")

            else:
                messagebox.showerror("Erro", f"Nenhum pedido encontrado referente ao nome {NOMEINPUT}")
                inputNota.delete(0, tk.END)
                inputNome.delete(0, tk.END)
                utils.focus(janelaBusca)
                inputNome.focus_set()
                return

        #busca pela nf
        elif NOTAINPUT != "":
            try:
                NOTAINPUT = int(NOTAINPUT)
                NOTAINPUT = str(NOTAINPUT)
            except:
                messagebox.showerror("ERRO", "Digite apenas numeros inteiros no campo de Nota!")
                inputNota.delete(0, tk.END)
                inputNome.delete(0, tk.END)
                utils.focus(janelaBusca)
                inputNota.focus_set()
                return

            NOTASENCONTRADAS = []
            for dia, listaPedidos in DB.items():
                for pedido in listaPedidos:
                    if pedido["nota"] == NOTAINPUT:
                        pedido["separado"] = dia
                        NOTASENCONTRADAS.append(pedido)

            if len(NOTASENCONTRADAS) > 0:
                LINHA = 0
                def on_mouse_wheel(event):
                    canvas.yview_scroll(-1 * (event.delta // 120), "units")

                #janela com a nf encontrada
                janelaResultado = tk.Toplevel(janela)
                janelaResultado.title('Notas Encontradas')

                frame = tk.Frame(janelaResultado)
                frame.grid(row=0, column=0, padx=5, pady=5, columnspan=2)

                tk.Label(frame, text="Notas Encontradas").grid(row=0, column=0, padx=5, pady=5, columnspan=2)

                canvas = tk.Canvas(frame, height=400)
                scrollbar = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
                scrollable_frame = tk.Frame(canvas)

                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(
                        scrollregion=canvas.bbox("all")
                    )
                )

                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)

                canvas.grid(row=1, column=0, sticky="nsew")
                scrollbar.grid(row=1, column=1, sticky="ns")

                janelaResultado.bind_all("<MouseWheel>", on_mouse_wheel)

                for nota in NOTASENCONTRADAS:
                    LINHA += 1
                    pedidoNota = nota["nota"]
                    pedidoNome = nota["nome"]
                    pedidoCidade = nota.get("cidade/ uf", " ")
                    pedidoVol = nota["vol."]
                    pedidoTrans = nota["transportadora"]
                    pedidoSep = nota.get("separador", " ")
                    pedidoRefri = nota.get("refrigerado", " ")
                    pedidoEnviado = nota["enviado"] 
                    dataSeparada = nota["separado"]

                    tk.Label(scrollable_frame, text=f"Nota: {pedidoNota}").grid(row=LINHA, column=0, padx=5, pady=5)
                    
                    listbox = tk.Listbox(scrollable_frame, height=7, width=40)
                    listbox.grid(row=LINHA, column=1, padx=5, pady=5)
                    listbox.insert(tk.END, f"Nome: {pedidoNome}")
                    listbox.insert(tk.END, f"Cidade/ UF: {pedidoCidade}")
                    listbox.insert(tk.END, f"Volumes: {pedidoVol}")
                    listbox.insert(tk.END, f"Refrigerado: {pedidoRefri}")
                    listbox.insert(tk.END, f"Transportadora: {pedidoTrans}")
                    listbox.insert(tk.END, f"Separado: {dataSeparada} por {pedidoSep}")
                    listbox.insert(tk.END, f"Enviado: {pedidoEnviado}")

            else:
                messagebox.showerror("Erro", f"Nenhum pedido encontrado referente a nota {NOTAINPUT}")
                inputNota.delete(0, tk.END)
                inputNome.delete(0, tk.END)
                utils.focus(janelaBusca)
                inputNota.focus_set()
                return

    #janela da busca onde sera digitado o nome do cliente ou o numero da nf
    janelaBusca=tk.Toplevel(janela)
    janelaBusca.title('Procurar Nota')

    tk.Label(janelaBusca,text='Nota').grid(row=0,column=0,padx=(10,5),pady=(5,0))
    inputNota=tk.Entry(janelaBusca, width=30)
    inputNota.grid(row=0,column=1,padx=(0,10),pady=(5,0))
    inputNota.focus_set()

    tk.Label(janelaBusca,text='Nome e Sobrenome').grid(row=1,column=0,padx=(10,5))
    inputNome=tk.Entry(janelaBusca, width=30)
    inputNome.grid(row=1,column=1,padx=(0,10))

    botaoProcurar= tk.Button(janelaBusca, text='Procurar', command=lambda: busca(), default="active", padx=15)
    botaoProcurar.grid(row=2,column=0,pady=(20,10),padx=10, sticky="w")

    janelaBusca.bind('<Return>', lambda event=None: botaoProcurar.invoke())

    botaoCancelar= tk.Button(janelaBusca, text='Cancelar', command=lambda:utils.cancelar(janelaBusca), padx=15)
    botaoCancelar.grid(row=2, column=1,pady=(20,10),padx=10, sticky="e")

#abre o excel com a separação do dia atual, ou do dia selecionado, ou ainda, de uma semana selecionada
def romaneioSeparacao(janela):
    utils.janelasKill(janela=janela)
    try:
        DB=carregarArquivo('DB.txt')
    except:
        messagebox.showerror('ERRO','Arquivo de dados inexistente!')
        utils.focus(janela=janela)
        return

    #função que procura pela data no DB
    def procurar(data):
        if data == "":
            return
        try:
            datetime.strptime(data, "%d/%m/%Y")
        except:
            messagebox.showerror("Data", "Data digitada inválida ou fora do padrão DD/MM/AAAA")
            procurarSemana.set(False)
            inputData.delete(0, tk.END)
            inputData.focus_set()
            return
        
        #função que vai criar e abrir o execel com o romaneio
        def criarExcel(semanal=bool, separadores=list, numTotalNotas=int, numTotalCxs=int, registros=list, datadigitada=None):
            wb = Workbook()
            ws = wb.active
            ws.page_setup.orientation = 'landscape'
            ws.title = "Romaneio de Separação"

            if semanal:
                ws["A1"]= f"Romaneio de Separação de {inicioSemana.strftime('%d/%m/%Y')} à {fimSemana.strftime('%d/%m/%Y')}"
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            else:
                ws["A1"]= f"Romaneio de Separação {datadigitada}"
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
            cell = ws.cell(row=1, column=1)
            cell.alignment = Alignment(horizontal="center")

            ws["A2"]=" "
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8 if semanal else 7)

            if semanal:
                argumentos = ["Nota", "Nome", "Cidade/ UF", "Vol.", "Transportadora", "Separador", "Separado", "Enviado"]
            else:
                argumentos = ["Nota", "Nome", "Cidade/ UF", "Vol.", "Transportadora", "Separador", "Enviado"]
            
            cabecalhos = list(argumentos)
            ws.append(cabecalhos)

            # Lista para armazenar o comprimento máximo de cada coluna
            maxLengths = [len(title) for title in cabecalhos]

            for registro in registros: 
                linha = []
                for cabecalho in cabecalhos:
                    valor = registro.get(cabecalho.lower(), "")

                    if cabecalho == "nome":
                        # Remove números no início do nome
                        valor = ''.join(filter(lambda x: not x.isdigit(), valor)).strip().lower().title()
                        
                        partesNome = valor.split()  # Divide o nome em partes

                        if len(partesNome) > 3: # Mantém no máximo três partes do nome
                            valor = ' '.join(partesNome[:3])

                    if cabecalho == "separador":
                        valor = valor.lower().strip().title()

                    linha.append(valor)

                # Atualiza o comprimento máximo das colunas baseado nos dados
                for i, cell_value in enumerate(linha):
                    # Calcula o comprimento máximo entre o título e o conteúdo
                    maxLengths[i] = max(maxLengths[i], len(str(cell_value)))

                ws.append(linha)
            
            # Ajustar o tamanho das colunas conforme o comprimento máximo
            for i, length in enumerate(maxLengths):
                adjustedWidth = length + 2  # Adiciona um pouco de espaço extra
                col_letter = chr(65 + i)  # Converte o índice da coluna para a letra (0 -> A, 1 -> B, etc.)
                ws.column_dimensions[col_letter].width = adjustedWidth
                
                # Alinhar o texto das células nas colunas
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=i+1)
                    # Alinhamento das células
                    cell.alignment = Alignment(horizontal="center")

            ultimaLinha = ws.max_row
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8 if semanal else 7):
                for cell in row:
                    cell.border = BORDAFINA

            ws.append([" "])
            ws.append([f"notas separadas: {numTotalNotas}"])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
            ws.append([f"caixas separadas: {numTotalCxs}"])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)

            for row in ws.iter_rows(min_row=ultimaLinha + 2, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = BORDAFINA

            ultimaLinha = ws.max_row

            ws.append([" "])
            ws.append(["Separadores:"])
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
            for sep in separadores:
                # Cada 'sep' é um dicionário com a estrutura {nome: {"notas separadas": n, "caixas separadas": c}}
                for nome, dados in sep.items():
                    ws.append([f"{nome} - Notas Separadas: {dados['notas separadas']}, Caixas Separadas: {dados['caixas separadas']}"])
                    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)

            for row in ws.iter_rows(min_row=ultimaLinha + 2, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = BORDAFINA

            wb.save("romaneio de separacao.xlsx")
            os.startfile("romaneio de separacao.xlsx")

        SEMANAL = False if not procurarSemana.get() else True
        ACHOU=False
        SEPARADORES = []
        NUMTOTALNOTAS = 0
        NUMTOTALCXS = 0
        if not SEMANAL:
            for dt in DB:
                if dt == data:
                    ACHOU=True
                    REGISTROS = DB.get(dt, [])
                    break
                    
        else:
            dia = datetime.strptime(data, "%d/%m/%Y")
            inicioSemana = dia - timedelta(days=dia.weekday())
            fimSemana = inicioSemana + timedelta(days=4)

            REGISTROS = []
            for chave, pedidos in DB.items():
                try:
                    dataChave = datetime.strptime(chave, "%d/%m/%Y")
                except ValueError:
                    # Se a data estiver em formato inesperado, pula
                    continue
                # Se a data do registro estiver dentro da semana
                if inicioSemana <= dataChave <= fimSemana:
                    for pedido in pedidos:
                        pedido["separado"] = chave
                    REGISTROS.extend(pedidos)
            
            if len(REGISTROS) >= 1:
                ACHOU = True

        if ACHOU == True:
            for pedido in REGISTROS:
                NUMTOTALCXS += int(pedido["vol."])
                NUMTOTALNOTAS += 1
                nomeSeparador = pedido["separador"].lower().strip().title()
                encontrado = False
                for item in SEPARADORES:
                    if nomeSeparador in item:
                        item[nomeSeparador]["notas separadas"] += 1
                        item[nomeSeparador]["caixas separadas"] += int(pedido["vol."])
                        encontrado = True
                        break
                if not encontrado:
                    SEPARADORES.append({nomeSeparador: {"notas separadas": 1,
                                                        "caixas separadas": int(pedido["vol."])}})

            criarExcel(semanal=SEMANAL, separadores=SEPARADORES, numTotalNotas=NUMTOTALNOTAS, numTotalCxs=NUMTOTALCXS, registros=REGISTROS, datadigitada=data if not SEMANAL else None)
            return

        else:
            messagebox.showerror('ERRO','Data não encontrada no DB')   
            utils.focus(janelaBuscaData)
            inputData.delete(0, tk.END)
            inputData.focus_set()
            return

    #janela do romaneio de separação
    janelaBuscaData=tk.Toplevel(janela)
    janelaBuscaData.title('Romaneio De Separação')

    tk.Label(janelaBuscaData,text='Data').grid(row=0,column=1,padx=5,pady=5)
    inputData=tk.Entry(janelaBuscaData, width=15)
    inputData.grid(row=1,column=1,padx=5,pady=5)
    inputData.focus_set()

    tk.Label(janelaBuscaData, text="DD/MM/AAAA").grid(row=2, column=1)

    #procura semanal ou so do dia selecionado
    procurarSemana = tk.BooleanVar()
    checkSemana = tk.Checkbutton(janelaBuscaData, text="Procura Semanal", variable=procurarSemana)
    checkSemana.grid(row=3, column=1, padx=5, pady=5)

    botaoProcurar= tk.Button(janelaBuscaData, text='Procurar', command=lambda: procurar(inputData.get().strip()), default="active", padx=15)
    botaoProcurar.grid(row=5,column=0,pady=(5,10),padx=10)

    #botao para data de hoje
    botaoHj=tk.Button(janelaBuscaData, text='Hoje', command=lambda: procurar(DATA), default="active", padx=25)
    botaoHj.grid(column=1, row=5, padx=10, pady=(5,10))

    botaoCancelar= tk.Button(janelaBuscaData, text='Cancelar', command=lambda:utils.cancelar(janelaBuscaData), padx=15)
    botaoCancelar.grid(row=5, column=2, pady=(5,10),padx=10) 

    janelaBuscaData.bind('<Return>', lambda event=None: botaoProcurar.invoke())
